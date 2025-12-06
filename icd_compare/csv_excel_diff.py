import argparse
import sys
import os
import csv
import tempfile
import atexit
import pandas as pd
import polars as pl
import xlsxwriter
from rapidfuzz import process, fuzz
import openpyxl
from openpyxl.utils import get_column_letter
import json

# Increase max string length for Polars just in case
pl.Config.set_fmt_str_lengths(1000)

# Global Debug Level
DEBUG_LEVEL = 0

# Track temporary files created when streaming Excel -> CSV so we can clean them up.
_TEMP_FILES = []


def _register_temp_file(path):
    _TEMP_FILES.append(path)


def _cleanup_temp_files():
    for p in _TEMP_FILES:
        try:
            os.remove(p)
        except Exception:
            pass


atexit.register(_cleanup_temp_files)

def log(msg, level=1):
    """
    Logs a message if DEBUG_LEVEL >= level.
    Level 1: Info / Progress
    Level 2: Flow / Arguments
    Level 3: Detailed Data / Inspection
    """
    if DEBUG_LEVEL >= level:
        prefix = "DEBUG" if level > 1 else "INFO"
        print(f"[{prefix}:{level}] {msg}")

def sanitize_for_excel(val):
    """Prevent formula injection in Excel."""
    if val is None:
        return ""
    s = str(val)
    if s.startswith(('=', '+', '-', '@')):
        return "'" + s
    return s


def _get_worksheet(wb, sheet_name):
    if sheet_name is None:
        return wb.active
    if isinstance(sheet_name, int):
        return wb.worksheets[sheet_name]
    return wb[sheet_name]


def stream_excel_to_temp_csv(path, header_row=1, sheet_name=None, max_rows=None):
    """
    Stream an Excel sheet to a temporary CSV using openpyxl read_only mode to avoid
    materializing the whole workbook in memory. Returns the temp CSV path.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = _get_worksheet(wb, sheet_name)
        header_idx = header_row - 1  # convert to 0-based
        rows_iter = ws.iter_rows(values_only=True)

        # Skip rows before the header row
        for _ in range(header_idx):
            next(rows_iter, None)

        header = next(rows_iter, None)
        if header is None:
            raise ValueError(f"Header row {header_row} not found in {path}")

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".csv", newline="", encoding="utf-8")
        _register_temp_file(tmp.name)
        writer = csv.writer(tmp, lineterminator="\n")

        normalized_header = [
            str(h) if h is not None else f"col_{idx+1}" for idx, h in enumerate(header)
        ]
        writer.writerow(normalized_header)

        written = 0
        for row in rows_iter:
            if max_rows is not None and written >= max_rows:
                break
            row_values = list(row[: len(normalized_header)])
            if len(row_values) < len(normalized_header):
                row_values.extend([""] * (len(normalized_header) - len(row_values)))
            writer.writerow([val if val is not None else "" for val in row_values])
            written += 1

        tmp.close()
        return tmp.name
    finally:
        try:
            wb.close()
        except Exception:
            pass


def read_excel_sample(path, header_row=1, sheet_name=None, n_rows=100):
    """
    Lightweight sampler for Excel: read header + up to n_rows of data via openpyxl streaming.
    Returns a Polars DataFrame.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = _get_worksheet(wb, sheet_name)
        header_idx = header_row - 1
        rows_iter = ws.iter_rows(values_only=True)
        for _ in range(header_idx):
            next(rows_iter, None)
        header = next(rows_iter, None)
        if header is None:
            raise ValueError(f"Header row {header_row} not found in {path}")

        normalized_header = [
            str(h) if h is not None else f"col_{idx+1}" for idx, h in enumerate(header)
        ]

        data = []
        for i, row in enumerate(rows_iter):
            if i >= n_rows:
                break
            row_values = list(row[: len(normalized_header)])
            if len(row_values) < len(normalized_header):
                row_values.extend([""] * (len(normalized_header) - len(row_values)))
            data.append(row_values)

        return pl.DataFrame(data, schema=normalized_header)
    finally:
        try:
            wb.close()
        except Exception:
            pass

def suggest_mapping(left_cols, right_cols):
    """
    Suggest mapping from left columns to right columns using fuzzy matching.
    Returns a dict: {left_col: suggested_right_col}
    """
    suggestions = {}
    # Create a pool of right columns to match against
    # We use a score cutoff to avoid bad matches
    for l in left_cols:
        # extractOne returns (match, score, index)
        match_result = process.extractOne(l, right_cols, scorer=fuzz.WRatio)
        if match_result:
            match, score, _ = match_result
            if score > 60: # Confidence threshold
                suggestions[l] = match
            else:
                suggestions[l] = ""
        else:
            suggestions[l] = ""
    return suggestions

def create_mapping_template(left_csv, right_csv, output_path, left_header_row=1, right_header_row=1, left_sheet=None, right_sheet=None):
    """
    Reads headers from CSVs and creates a mapping template Excel file.
    """
    log(f"Reading headers from {left_csv} and {right_csv}...", 1)
    try:
        # Read only headers
        l_df = read_data_eager_headers(left_csv, header_row=left_header_row, sheet_name=left_sheet)
        r_df = read_data_eager_headers(right_csv, header_row=right_header_row, sheet_name=right_sheet)
        left_cols = l_df.columns
        right_cols = r_df.columns
    except Exception as e:
        log(f"Error reading CSV headers: {e}", 1)
        sys.exit(1)

    log("Generating suggestions...", 1)
    suggestions = suggest_mapping(left_cols, right_cols)

    # Smart Fill Detection
    # Read a sample of Left file to detect fill_down candidates
    # Heuristic: Valid data exists, followed by gaps.
    log("Analyzing left file for Fill Down candidates...", 1)
    fill_down_suggestions = {}
    try:
        # Read up to 100 rows without loading the whole Excel into memory
        if str(left_csv).lower().endswith(('.xlsx', '.xls')):
            df_sample = read_excel_sample(left_csv, header_row=left_header_row, sheet_name=left_sheet, n_rows=100)
        else:
            df_sample = read_data_lazy(left_csv, header_row=left_header_row, sheet_name=left_sheet).head(100).collect()
        
        log(f"Sample data columns: {df_sample.columns}", 3)
        
        for col in left_cols:
            if col in df_sample.columns:
                s = df_sample[col]
                
                # Check for ANY valid data first
                # We need to find the first index that is valid
                is_null_mask = s.is_null() | s.cast(pl.String).str.strip_chars().eq("")
                # Get valid indices
                valid_indices = is_null_mask.not_().arg_true()
                
                if not valid_indices.is_empty():
                    first_valid_idx = valid_indices[0]
                    # Check if there are gaps AFTER the first valid index
                    # We look at the slice from first_valid_idx to end
                    s_tail = s.slice(first_valid_idx + 1)
                    
                    if not s_tail.is_empty():
                        tail_nulls = s_tail.is_null() | s_tail.cast(pl.String).str.strip_chars().eq("")
                        has_gaps = tail_nulls.any()
                        
                        if has_gaps:
                            fill_down_suggestions[col] = "Y"
                            log(f"Suggesting Fill Down for '{col}' (First valid at {first_valid_idx}, found gaps)", 2)
                        else:
                            fill_down_suggestions[col] = ""
                    else:
                         fill_down_suggestions[col] = ""
                else:
                    # No valid data at all
                     fill_down_suggestions[col] = ""

    except Exception as e:
        log(f"Warning: Could not analyze for Fill Down suggestions: {e}", 1)

    # Create DataFrame for the mapping sheet
    mapping_data = []
    for l in left_cols:
        mapping_data.append({
            "left_column": l,
            "suggested_right_column": suggestions.get(l, ""),
            "confirmed_right_column": suggestions.get(l, ""), # Default to suggestion
            "is_key": "",
            "fill_down": fill_down_suggestions.get(l, "")
        })
    
    df_mapping = pd.DataFrame(mapping_data)

    # Create Instructions DataFrame
    instructions_data = [
        ["Step 1", "Review the 'Columns' sheet."],
        ["Step 2", "Verify 'confirmed_right_column' matches the correct column in the Right file. Clear it if you don't want to compare that column."],
        ["Step 3", "Mark Key columns by entering 'Y' in 'is_key'. Keys are used to join rows."],
        ["Step 4", "Mark columns that need Forward Fill (e.g. parent IDs) by entering 'Y' in 'fill_down'. This fills empty cells with the value from the row above."],
        ["Step 5", "Save this workbook."],
        ["Step 6", "Run the diff script again with --mapping-confirmed."]
    ]
    df_instructions = pd.DataFrame(instructions_data, columns=["Step", "Action"])

    print(f"Writing mapping template to {output_path}...")
    
    path_str = str(output_path).lower()
    
    # CASE 1: CSV Output
    if path_str.endswith('.csv'):
        # Just write the mapping dataframe
        try:
             df_mapping.to_csv(output_path, index=False)
             print(f"Created CSV mapping template at {output_path}")
             print("Note: CSV format does not support dropdowns or instructions sheet.")
             print("Usage: Edit the CSV directly. Set 'is_key' and 'fill_down' to 'Y'.")
             return
        except Exception as e:
             print(f"Error writing CSV template: {e}")
             sys.exit(1)
             
    # CASE 2: Excel Output
    with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
        df_mapping.to_excel(writer, sheet_name='Columns', index=False)
        df_instructions.to_excel(writer, sheet_name='Instructions', index=False)
        
        # Add some formatting
        workbook = writer.book
        worksheet = writer.sheets['Columns']
        
        # Add a dropdown for is_key and fill_down
        validation_yn = {'validate': 'list', 'source': ['Y', 'N']}
        worksheet.data_validation(f'D2:D{len(df_mapping)+1}', validation_yn) 
        worksheet.data_validation(f'E2:E{len(df_mapping)+1}', validation_yn) 
        
        # Add a dropdown for confirmed_right_column
        # We need to list all available right columns
        # Excel validation list has a limit of 255 chars if passed directly.
        # Better to write the list to a hidden sheet and reference it.
        
        # Create a hidden sheet for validation lists
        worksheet_lists = workbook.add_worksheet('ValidationLists')
        worksheet_lists.hide()
        worksheet_lists.write_column('A1', right_cols)
        
        # Define the range for the list
        right_cols_len = len(right_cols)
        if right_cols_len > 0:
            list_formula = f'=ValidationLists!$A$1:$A${right_cols_len}'
            validation_right = {'validate': 'list', 'source': list_formula}
            worksheet.data_validation(f'C2:C{len(df_mapping)+1}', validation_right)

        # Adjust column widths
        worksheet.set_column('A:A', 30)
        worksheet.set_column('B:B', 30)
        worksheet.set_column('C:C', 30)
        worksheet.set_column('D:D', 10)
        worksheet.set_column('E:E', 15)

    print(f"Done. Please edit the mapping file and re-run.")

def read_mapping(mapping_path):
    """
    Reads the confirmed mapping file using openpyxl (read-only) for memory safety.
    Returns:
        mapping_dict: {left_col: right_col} (only for confirmed mappings)
        keys: list of left_col names that are keys
        fill_down_cols: list of left_col names to forward fill
    """
    print(f"Reading mapping from {mapping_path}...")
    
    mapping_dict = {}
    keys = []
    fill_down_cols = []
    
    path_str = str(mapping_path).lower()
    
    # CASE 1: CSV (Simpler, Robust)
    if path_str.endswith('.csv'):
        try:
            df = pd.read_csv(mapping_path)
            # Normalize
            df.columns = [c.lower().strip() for c in df.columns]
            
            required = {'left_column', 'confirmed_right_column', 'is_key'}
            if not required.issubset(df.columns):
                 print(f"Error: CSV mapping missing required columns: {required}")
                 sys.exit(1)
                 
            for _, row in df.iterrows():
                l_col = str(row['left_column']).strip()
                r_col = str(row['confirmed_right_column']).strip() # Handle NaN
                if pd.isna(row['confirmed_right_column']): r_col = ""
                
                is_key = str(row['is_key']).strip().upper() == 'Y'
                fill_down = False
                if 'fill_down' in df.columns:
                     fill_down = str(row['fill_down']).strip().upper() == 'Y'
                     
                if r_col and r_col.lower() != 'nan':
                     mapping_dict[l_col] = r_col
                     if is_key: keys.append(l_col)
                     if fill_down: fill_down_cols.append(l_col)
                elif is_key:
                     print(f"Warning: Column '{l_col}' is marked as key but has no mapped right column. Ignoring.")
                     
            return mapping_dict, keys, fill_down_cols
            
        except Exception as e:
            print(f"Error reading CSV mapping: {e}")
            sys.exit(1)

    # CASE 2: EXCEL (Original Logic)
    try:
        # Use read_only=True to handle massive files (phantom rows) without OOM
        wb = openpyxl.load_workbook(mapping_path, read_only=True, data_only=True)
        if 'Columns' not in wb.sheetnames:
             print(f"Error: Mapping file missing 'Columns' sheet.")
             sys.exit(1)
        ws = wb['Columns']
    except Exception as e:
        print(f"Error reading mapping file: {e}")
        sys.exit(1)
    
    
    # Header Mapping
    # expected headers: left_column, confirmed_right_column, is_key, fill_down
    # We need to find the specific column indices.
    header_map = {}
    headers_found = False
    
    # Iterator
    row_iter = ws.iter_rows(values_only=True)
    
    # 1. Find Headers
    for row in row_iter:
        # Check if this row looks like a header
        # We look for 'left_column'
        row_lower = [str(c).strip().lower() if c is not None else "" for c in row]
        if 'left_column' in row_lower:
            # Found headers
            for idx, val in enumerate(row_lower):
                if val:
                    header_map[val] = idx
            headers_found = True
            break
            
    if not headers_found:
        print("Error: Could not find 'left_column' header in 'Columns' sheet.")
        sys.exit(1)

    required = {'left_column', 'confirmed_right_column', 'is_key'}
    if not required.issubset(header_map.keys()):
        print(f"Error: Mapping sheet missing required columns: {required}")
        sys.exit(1)
        
    col_idx_left = header_map['left_column']
    col_idx_right = header_map['confirmed_right_column']
    col_idx_key = header_map['is_key']
    col_idx_fill = header_map.get('fill_down') # Optional
    
    # 2. Iterate Data
    # Stop if we encounter too many consecutive empty rows (phantom row protection)
    consecutive_empty = 0
    max_empty = 20
    
    for row in row_iter:
        # Row values tuple
        
        # Check if row is empty
        if not any(row):
            consecutive_empty += 1
            if consecutive_empty > max_empty:
                break
            continue
        consecutive_empty = 0 # Reset if valid data found
        
        # Get values safely
        v_left = row[col_idx_left]
        v_right = row[col_idx_right]
        v_key = row[col_idx_key]
        v_fill = row[col_idx_fill] if col_idx_fill is not None else None
        
        # Process left column
        l_col = str(v_left).strip() if v_left is not None else ""
        if not l_col:
            continue
            
        # Process right column
        r_col = str(v_right).strip() if v_right is not None else ""
        
        # Process flags
        is_key = str(v_key).strip().upper() == 'Y' if v_key is not None else False
        fill_down = str(v_fill).strip().upper() == 'Y' if v_fill is not None else False
        
        if r_col:
            mapping_dict[l_col] = r_col
            if is_key:
                keys.append(l_col)
            if fill_down:
                fill_down_cols.append(l_col)
        elif is_key:
             print(f"Warning: Column '{l_col}' is marked as key but has no mapped right column. Ignoring.")
             
    try:
        wb.close()
    except:
        pass
        
    return mapping_dict, keys, fill_down_cols

def read_data_lazy(path, header_row=1, sheet_name=None):
    """
    Reads data lazily from CSV or Excel, respecting the header_row (1-based).
    """
    path_str = str(path).lower()
    # 0-based index for polars/pandas
    header_idx = header_row - 1
    
    log(f"Reading data lazy: {path} (Sheet: {sheet_name}, Header: {header_row})", 2)
    
    if path_str.endswith('.csv'):
        # Polars scan_csv uses skip_rows to skip lines BEFORE the header
        # If header is on row 1, skip_rows=0. If row 2, skip_rows=1.
        return pl.scan_csv(path, skip_rows=header_idx)
    elif path_str.endswith(('.xlsx', '.xls')):
        # Stream Excel -> temp CSV to avoid loading the entire workbook into memory
        try:
            temp_csv = stream_excel_to_temp_csv(path, header_row=header_row, sheet_name=sheet_name)
            # Header already included; no skip_rows needed.
            return pl.scan_csv(temp_csv)
        except Exception as e:
            raise ValueError(f"Error streaming Excel file {path}: {e}")
    else:
        raise ValueError(f"Unsupported file format: {path}")

def read_data_eager_headers(path, header_row=1, sheet_name=None):
    """
    Reads headers eagerly for mapping generation.
    """
    path_str = str(path).lower()
    header_idx = header_row - 1
    
    log(f"Reading headers eager: {path} (Sheet: {sheet_name}, Header: {header_row})", 2)

    if path_str.endswith('.csv'):
        # read_csv also supports skip_rows
        return pl.read_csv(path, n_rows=0, skip_rows=header_idx)
    elif path_str.endswith(('.xlsx', '.xls')):
        try:
             log(f"Reading Excel Headers '{path}' with header={header_idx}, sheet_name={sheet_name}", 3)
             temp_csv = stream_excel_to_temp_csv(path, header_row=header_row, sheet_name=sheet_name, max_rows=0)
             return pl.read_csv(temp_csv, n_rows=0)
        except Exception as e:
             raise ValueError(f"Error reading Excel headers {path}: {e}")
    else:
        raise ValueError(f"Unsupported file format: {path}")



def compute_diff(left_path, right_path, mapping, keys, fill_down_cols=None, left_header_row=1, right_header_row=1, left_sheet=None, right_sheet=None):
    """
    Computes the diff using Polars.
    """
    log("Loading data into Polars...", 1)
    log(f"  Left: {left_path} (Header: {left_header_row}, Sheet: {left_sheet})", 1)
    log(f"  Right: {right_path} (Header: {right_header_row}, Sheet: {right_sheet})", 1)
    
    # Scan Data (lazy)
    lf_left = read_data_lazy(left_path, header_row=left_header_row, sheet_name=left_sheet)
    lf_right = read_data_lazy(right_path, header_row=right_header_row, sheet_name=right_sheet)

    # 0. Apply Forward Fill if requested
    # We do this BEFORE any renaming or joining.
    # For Left:
    if fill_down_cols:
        log(f"Applying Forward Fill to: {fill_down_cols}", 2)
        # Only apply to cols present in Left
        # Note: Polars forward_fill works on Nulls. 
        # If the file has empty strings, we must replace "" with Null first.
        
        # Helper to setup fill exprs
        def diff_fill(lf, cols_to_fill):
            # 1. Replace empty strings with null
            # 2. Forward fill
            # We assume cols_to_fill exist.
            exprs = []
            for c in cols_to_fill:
                 exprs.append(
                     pl.when(pl.col(c).cast(pl.String).str.strip_chars() == "")
                     .then(None)
                     .otherwise(pl.col(c))
                     .forward_fill()
                     .alias(c)
                 )
            return lf.with_columns(exprs)

        # Apply to Left
        # We need to intersect with available columns in lazyframe?
        # Actually mapping keys are Left columns. So fill_down_cols are Left columns.
        lf_left = diff_fill(lf_left, [c for c in fill_down_cols if c in mapping]) # simplistic check

        # Apply to Right
        # Right columns have different names (Values of mapping).
        # We need to map `fill_down_cols` (which are Left names) to Right names.
        right_fill_cols = [mapping[c] for c in fill_down_cols if c in mapping]
        lf_right = diff_fill(lf_right, right_fill_cols)

    # Select and rename columns in Right to match Left (for keys and mapped cols)
    # Strategy: 
    # 1. Rename Right columns to their Left counterparts (so we can join on keys).
    # 2. But wait, if we join on keys, non-key columns will collide.
    #    Polars join suffix defaults to "_right".
    #    So if we rename R_Col -> L_Col, and join, we get L_Col (from Left) and L_Col_right (from Right).
    #    This is exactly what we want.
    
    right_selects = []
    for l_col, r_col in mapping.items():
        right_selects.append(pl.col(r_col).alias(l_col))
    
    # Also keep unmapped columns from right? 
    # The requirement says "Left_only sheet: rows present only in left CSV", "Right_only sheet: rows present only in right CSV".
    # Usually we want to see the original data.
    # But for the diff logic, we only care about mapped columns.
    # Let's keep all columns for the "Left/Right Only" exports, but for the "Changed" logic we focus on mapped.
    
    # Actually, to produce "Right Only" rows with original schema, we need the original right columns.
    # But if we rename them, we lose them?
    # We can keep them or reconstruct them.
    # Let's just select the mapped ones for the diffing dataframe.
    # We can reload or join back for full details if needed, but for now let's assume we just want mapped columns in the diff report?
    # Requirement: "Right_only sheet: rows present only in right CSV."
    # If we only select mapped columns, we lose unmapped ones.
    # Let's try to keep everything but aliased.
    
    # Simplified approach:
    # 1. Rename mapped right columns to `L_Col` (so they align with Left).
    # 2. Join.
    # 3. Mapped columns will be `L_Col` and `L_Col_right`.
    # 4. Unmapped right columns? If we rename `R_Col` -> `L_Col`, `R_Col` is gone.
    #    If `R_Col` was NOT mapped, it stays `R_Col`.
    #    But if `R_Col` collides with a `L_Col` (unlikely but possible), it might be confusing.
    #    Let's assume mapping is authoritative.
    
    # Construct the rename mapping for right
    right_rename_map = {v: k for k, v in mapping.items()} # right_name -> left_name
    
    # Apply renaming to right lazyframe
    # We use select to reorder/rename, but we want to keep unmapped ones too?
    # It's easier to just rename.
    lf_right_renamed = lf_right.rename(right_rename_map)

    # Let's add the flags *before* the join in Step 0.
    # But I can't restart Step 0 easily here without rewriting everything.
    # Alternative: check if key columns are null.
    # If Join Key is present, the row exists.
    # But keys can be null in data? (Not if we fill down properly).
    
    # Let's stick to the separation for now but apply fill logic to the marking frames too?
    # Or better: Use the `joined` dataframe and verify presence by checking non-null keys?
    # If we assume keys are populated (especially with fill-down), checking keys is safe.
    # BUT, let's keep the explicit flags to remain robust against null keys.
    
    # So, I need to apply Fill Down to the marked frames too.
    # This suggests I should refactor `read_and_process` or just copy logic.
    # To keep it simple:
    lf_left_marked = read_data_lazy(left_path, header_row=left_header_row, sheet_name=left_sheet)
    if fill_down_cols:
         lf_left_marked = diff_fill(lf_left_marked, [c for c in fill_down_cols if c in mapping])
    lf_left_marked = lf_left_marked.with_columns(pl.lit(True).alias("_in_left"))
    
    lf_right_marked = read_data_lazy(right_path, header_row=right_header_row, sheet_name=right_sheet)
    if fill_down_cols:
         lf_right_marked = diff_fill(lf_right_marked, right_fill_cols)
    lf_right_marked = lf_right_marked.rename(right_rename_map).with_columns(pl.lit(True).alias("_in_right"))
    
    # If no keys, we need a fallback.
    if not keys:
        log("No keys provided. Using full-row hash comparison.", 1)
        # Create a hash column for both
        # We can only hash mapped columns to ensure comparison is valid
        mapped_l_cols = list(mapping.keys())
        
        lf_left_marked = lf_left_marked.with_columns(
            pl.concat_str(pl.col(mapped_l_cols)).hash().alias("_row_hash")
        )
        lf_right_marked = lf_right_marked.with_columns(
            pl.concat_str(pl.col(mapped_l_cols)).hash().alias("_row_hash")
        )
        keys = ["_row_hash"]

    log(f"Joining on keys: {keys}", 1)

    joined = lf_left_marked.join(lf_right_marked, on=keys, how="full", suffix="_right")
    df = joined.collect()
    
    # Fill nulls in presence flags
    df = df.with_columns([
        pl.col("_in_left").fill_null(False),
        pl.col("_in_right").fill_null(False)
    ])
    
    # Define Merge Status
    # Left Only: _in_left & ~_in_right
    # Right Only: ~_in_left & _in_right
    # Both: _in_left & _in_right
    
    # Identify Changed Rows
    # For rows in 'Both', compare mapped columns.
    # Mapped columns: `col` vs `col_right`
    
    # Build expressions for differences
    diff_exprs = []
    changed_col_names_expr = pl.lit("") # Start with empty string
    
    # We only compare columns that are NOT keys (keys are by definition equal in a join match)
    compare_cols = [c for c in mapping.keys() if c not in keys]
    
    # We need to calculate 'is_changed' for the row
    # And also collect a list of changed column names
    
    # Let's do this efficiently.
    # We can create a boolean column for each comparison
    # Then aggregate them.
    
    check_cols = []
    for col in compare_cols:
        right_col = f"{col}_right"
        # Compare col vs right_col
        # Use eq_missing to handle nulls safely (null==null is True)
        # We want difference, so .not_()
        is_diff = pl.col(col).eq_missing(pl.col(right_col)).not_()
        
        # We only care if it's a diff AND it's in both (otherwise it's just missing data)
        # But if it's Left Only or Right Only, we don't flag as "Changed" in the diff sense.
        is_diff = is_diff & pl.col("_in_left") & pl.col("_in_right")
        
        check_cols.append(is_diff.alias(f"_diff_{col}"))

    if check_cols:
        df = df.with_columns(check_cols)
        
        # Now create the comma-separated list of changed columns
        # This is a bit tricky in Polars without fold/concat_str on conditions
        # We can use `pl.concat_str` with `pl.when`
        
        concat_exprs = []
        for col in compare_cols:
            # If diff, add "colname, ", else ""
            concat_exprs.append(
                pl.when(pl.col(f"_diff_{col}"))
                .then(pl.lit(col + ", "))
                .otherwise(pl.lit(""))
            )
        
        df = df.with_columns(
            pl.concat_str(concat_exprs).str.strip_chars(", ").alias("changed_columns")
        )
        
        # Determine final status
        df = df.with_columns(
            pl.when(pl.col("_in_left") & pl.col("_in_right") & (pl.col("changed_columns") != ""))
            .then(pl.lit("changed"))
            .when(pl.col("_in_left") & pl.col("_in_right"))
            .then(pl.lit("equal"))
            .when(pl.col("_in_left"))
            .then(pl.lit("left_only"))
            .otherwise(pl.lit("right_only"))
            .alias("_merge")
        )
    else:
        # No columns to compare (only keys?)
        df = df.with_columns(
             pl.when(pl.col("_in_left") & pl.col("_in_right"))
            .then(pl.lit("equal"))
            .when(pl.col("_in_left"))
            .then(pl.lit("left_only"))
            .otherwise(pl.lit("right_only"))
            .alias("_merge")
        ).with_columns(pl.lit("").alias("changed_columns"))

    return df

def write_excel_report(df, output_path, mapping, keys, max_rows=200000):
    """
    Writes the rich Excel report.
    """
    log(f"Writing Excel report to {output_path}...", 1)
    
    # Split data
    left_only = df.filter(pl.col("_merge") == "left_only")
    right_only = df.filter(pl.col("_merge") == "right_only")
    changed = df.filter(pl.col("_merge") == "changed")
    
    # Counts
    counts = {
        "Total Rows": len(df),
        "Left Only": len(left_only),
        "Right Only": len(right_only),
        "Changed": len(changed),
        "Equal": len(df) - len(left_only) - len(right_only) - len(changed)
    }
    
    # Prepare Pandas DataFrames for export (limit rows if needed)
    # Helper to convert and truncate
    def to_pandas(pl_df, limit):
        if len(pl_df) > limit:
            log(f"Warning: Truncating sheet to {limit} rows (Total: {len(pl_df)})", 1)
            return pl_df.head(limit).to_pandas()
        return pl_df.to_pandas()

    # Create Excel Writer
    writer = pd.ExcelWriter(output_path, engine='xlsxwriter')
    workbook = writer.book
    
    # Formats
    fmt_header = workbook.add_format({'bold': True, 'bg_color': '#D3D3D3', 'border': 1})
    fmt_red_strike = workbook.add_format({'font_color': '#C9302C', 'font_strikeout': True})
    fmt_blue = workbook.add_format({'font_color': '#0275D8'})
    fmt_arrow = workbook.add_format({'font_color': '#000000'}) # Neutral
    fmt_changed_row = workbook.add_format({'bg_color': '#FFF3CD'}) # Light orange
    
    # 1. Summary Sheet
    summary_df = pd.DataFrame(list(counts.items()), columns=["Metric", "Count"])
    summary_df.to_excel(writer, sheet_name="Summary", index=False)
    ws_summary = writer.sheets["Summary"]
    ws_summary.set_column("A:A", 20)
    ws_summary.set_column("B:B", 15)
    
    # Add links
    ws_summary.write_url("D2", "internal:'Left_only'!A1", string="Go to Left Only")
    ws_summary.write_url("D3", "internal:'Right_only'!A1", string="Go to Right Only")
    ws_summary.write_url("D4", "internal:'Changed_rich'!A1", string="Go to Changed (Rich)")
    ws_summary.write_url("D5", "internal:'Changed_side_by_side'!A1", string="Go to Changed (Side-by-Side)")

    # 2. Left Only & Right Only
    to_pandas(left_only, max_rows).to_excel(writer, sheet_name="Left_only", index=False)
    to_pandas(right_only, max_rows).to_excel(writer, sheet_name="Right_only", index=False)
    
    # 3. Changed Rich
    pdf_changed = to_pandas(changed, max_rows)
    
    if not pdf_changed.empty:
        # Create sheet
        ws_rich = workbook.add_worksheet("Changed_rich")
        writer.sheets["Changed_rich"] = ws_rich
        
        # Write Header
        display_cols = keys + [c for c in mapping.keys() if c not in keys]
        
        for col_idx, col_name in enumerate(display_cols):
            ws_rich.write(0, col_idx, col_name, fmt_header)
            
        # Write Rows
        for row_idx, row in pdf_changed.iterrows():
            excel_row = row_idx + 1
            
            # Check which columns changed
            changed_cols_str = str(row.get("changed_columns", ""))
            changed_set = set(c.strip() for c in changed_cols_str.split(",") if c.strip())
            
            for col_idx, col_name in enumerate(display_cols):
                val_left = row.get(col_name)
                val_right = row.get(f"{col_name}_right")
                
                if col_name in keys:
                    ws_rich.write(excel_row, col_idx, sanitize_for_excel(val_left))
                elif col_name in changed_set:
                    # Changed: Write Rich String
                    s_old = str(val_left) if pd.notna(val_left) else "NULL"
                    s_new = str(val_right) if pd.notna(val_right) else "NULL"
                    
                    try:
                        ws_rich.write_rich_string(excel_row, col_idx,
                                                fmt_red_strike, s_old,
                                                fmt_arrow, " -> ",
                                                fmt_blue, s_new)
                    except Exception:
                        ws_rich.write(excel_row, col_idx, f"{s_old} -> {s_new}")
                else:
                    ws_rich.write(excel_row, col_idx, sanitize_for_excel(val_left))

    # 4. Changed Side by Side
    if not pdf_changed.empty:
        ws_sbs = workbook.add_worksheet("Changed_side_by_side")
        writer.sheets["Changed_side_by_side"] = ws_sbs
        
        # Header
        headers = []
        for k in keys:
            headers.append(k)
        for c in mapping.keys():
            if c not in keys:
                headers.append(f"{c}_old")
                headers.append(f"{c}_new")
        
        for i, h in enumerate(headers):
            ws_sbs.write(0, i, h, fmt_header)
            
        # Rows
        for row_idx, row in pdf_changed.iterrows():
            excel_row = row_idx + 1
            col_ptr = 0
            
            # Keys
            for k in keys:
                ws_sbs.write(excel_row, col_ptr, sanitize_for_excel(row.get(k)))
                col_ptr += 1
                
            # Values
            for c in mapping.keys():
                if c not in keys:
                    val_left = row.get(c)
                    val_right = row.get(f"{c}_right")
                    
                    ws_sbs.write(excel_row, col_ptr, sanitize_for_excel(val_left))
                    col_ptr += 1
                    
                    ws_sbs.write(excel_row, col_ptr, sanitize_for_excel(val_right))
                    col_ptr += 1

    writer.close()
    log("Excel report generated successfully.", 1)

def write_html_report(df, output_path, mapping, keys, hierarchy_cols):
    """
    Writes a rich HTML report with hierarchical navigation.
    """
    log(f"Writing HTML report to {output_path}...", 1)
    
    # Convert to pandas for easier HTML generation (assuming it fits in memory)
    pdf = df.to_pandas()
    
    # Pre-process for hierarchy
    # We need to build a tree structure: Level 1 -> Level 2 -> [Rows]
    
    # 1. Build Tree Data
    tree_data = {}
    
    # Helper to insert into tree
    def insert_tree(node, path, row_index, status):
        if not path:
            if "_rows" not in node:
                node["_rows"] = []
            node["_rows"].append({"idx": row_index, "status": status})
            return
        
        head = path[0]
        tail = path[1:]
        if head not in node:
            node[head] = {}
        insert_tree(node[head], tail, row_index, status)

    # Let's add a 'status' column for easier processing
    pdf['status'] = pdf['_merge']
    
    # Prepare data for JSON
    display_cols = keys + [k for k in mapping.keys() if k not in keys]
    
    # Create a list of dicts for the data
    data_list = []
    for idx, row in pdf.iterrows():
        row_data = {"_id": idx, "status": row['status']}
        for col in display_cols:
            val_left = row.get(col)
            val_right = row.get(f"{col}_right")
            
            # Formatting
            if row['status'] == 'changed':
                changed_cols = str(row.get("changed_columns", "")).split(",")
                changed_cols = [c.strip() for c in changed_cols]
                
                if col in changed_cols:
                    row_data[col] = {
                        "val": f'<span class="diff-old">{val_left}</span> <span class="diff-arrow">&rarr;</span> <span class="diff-new">{val_right}</span>',
                        "is_diff": True
                    }
                else:
                    row_data[col] = {"val": val_left, "is_diff": False}
            elif row['status'] == 'left_only':
                 row_data[col] = {"val": val_left, "is_diff": False}
            elif row['status'] == 'right_only':
                 row_data[col] = {"val": val_right, "is_diff": False}
            else: # same
                 row_data[col] = {"val": val_left, "is_diff": False}
        
        # Add hierarchy values for the tree
        h_values = []
        for h_col in hierarchy_cols:
            val = row.get(h_col)
            if pd.isna(val) and f"{h_col}_right" in row:
                val = row.get(f"{h_col}_right")
            h_values.append(str(val) if pd.notna(val) else "N/A")
        
        row_data["_h_values"] = h_values
        data_list.append(row_data)

    # Build the tree structure for the sidebar
    tree_root = {}
    
    def add_to_tree(node, path, row_idx, status):
        if "_stats" not in node:
            node["_stats"] = {"total": 0, "changed": 0, "left_only": 0, "right_only": 0, "same": 0}
        
        node["_stats"]["total"] += 1
        if status in node["_stats"]:
            node["_stats"][status] += 1
            
        if not path:
            if "_ids" not in node:
                node["_ids"] = []
            node["_ids"].append(row_idx)
            return

        head = path[0]
        tail = path[1:]
        if head not in node:
            node[head] = {}
        add_to_tree(node[head], tail, row_idx, status)

    for i, item in enumerate(data_list):
        add_to_tree(tree_root, item["_h_values"], i, item["status"])

    # HTML Template (Shortened for brevity here, assuming it works)
    # I'll rely on the existing template logic from temp_restore.py but minified slightly or just copied.
    # To save space in this response, I'll trust the previous content was fine.
    # But wait, I need to WRITE it.
    
    html_content = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>ICD Diff Report</title>
    <style>
        :root {{ --primary-color: #2c3e50; --secondary-color: #34495e; --accent-color: #3498db; --text-color: #333; --bg-color: #f4f6f7; --sidebar-width: 300px; }}
        body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; margin: 0; display: flex; height: 100vh; background-color: var(--bg-color); color: var(--text-color); }}
        #sidebar {{ width: var(--sidebar-width); background-color: white; border-right: 1px solid #ddd; overflow-y: auto; padding: 20px; box-shadow: 2px 0 5px rgba(0,0,0,0.05); }}
        #main {{ flex: 1; padding: 20px; overflow-y: auto; display: flex; flex-direction: column; }}
        h1, h2, h3 {{ color: var(--primary-color); }}
        .tree-node {{ margin-left: 15px; }}
        .tree-label {{ cursor: pointer; padding: 5px; border-radius: 4px; display: flex; justify-content: space-between; align-items: center; }}
        .tree-label:hover {{ background-color: #eef2f7; }}
        .tree-label.active {{ background-color: var(--accent-color); color: white; }}
        .tree-children {{ display: none; }}
        .tree-children.open {{ display: block; }}
        .badge {{ font-size: 0.8em; padding: 2px 6px; border-radius: 10px; background-color: #eee; color: #666; }}
        .badge.changed {{ background-color: #fff3cd; color: #856404; }}
        table {{ width: 100%; border-collapse: collapse; background: white; box-shadow: 0 1px 3px rgba(0,0,0,0.1); margin-top: 20px; }}
        th, td {{ padding: 12px 15px; text-align: left; border-bottom: 1px solid #eee; }}
        th {{ background-color: var(--secondary-color); color: white; position: sticky; top: 0; }}
        tr:hover {{ background-color: #f9f9f9; }}
        .diff-old {{ color: #e74c3c; text-decoration: line-through; }}
        .diff-new {{ color: #27ae60; font-weight: bold; }}
        .diff-arrow {{ color: #95a5a6; margin: 0 5px; }}
        .status-dot {{ height: 10px; width: 10px; border-radius: 50%; display: inline-block; margin-right: 5px; }}
        .status-changed {{ background-color: #f1c40f; }}
        .status-left-only {{ background-color: #e74c3c; }}
        .status-right-only {{ background-color: #2ecc71; }}
        .status-same {{ background-color: #bdc3c7; }}
        .summary-cards {{ display: flex; gap: 20px; margin-bottom: 20px; }}
        .card {{ background: white; padding: 15px; border-radius: 8px; box-shadow: 0 2px 5px rgba(0,0,0,0.05); flex: 1; text-align: center; }}
        .card-number {{ font-size: 2em; font-weight: bold; color: var(--accent-color); }}
        .card-label {{ color: #7f8c8d; }}
    </style>
</head>
<body>
<div id="sidebar"><h3>Navigation</h3><div id="tree-root"></div></div>
<div id="main">
    <div class="summary-cards">
        <div class="card"><div class="card-number" id="count-total">0</div><div class="card-label">Total Rows</div></div>
        <div class="card"><div class="card-number" id="count-changed" style="color: #f1c40f">0</div><div class="card-label">Changed</div></div>
        <div class="card"><div class="card-number" id="count-left" style="color: #e74c3c">0</div><div class="card-label">Left Only</div></div>
        <div class="card"><div class="card-number" id="count-right" style="color: #2ecc71">0</div><div class="card-label">Right Only</div></div>
    </div>
    <div id="table-container">
        <h2 id="current-view-title">All Data</h2>
        <table id="data-table"><thead><tr><th>Status</th>{''.join(f'<th>{col}</th>' for col in display_cols)}</tr></thead><tbody></tbody></table>
    </div>
</div>
<script>
    const data = {json.dumps(data_list)};
    const tree = {json.dumps(tree_root)};
    const displayCols = {json.dumps(display_cols)};
    document.getElementById('count-total').innerText = data.length;
    document.getElementById('count-changed').innerText = data.filter(r => r.status === 'changed').length;
    document.getElementById('count-left').innerText = data.filter(r => r.status === 'left_only').length;
    document.getElementById('count-right').innerText = data.filter(r => r.status === 'right_only').length;
    function renderTable(rows) {{
        const tbody = document.querySelector('#data-table tbody');
        tbody.innerHTML = '';
        rows.forEach(row => {{
            const tr = document.createElement('tr');
            const statusTd = document.createElement('td');
            const dot = document.createElement('span');
            dot.className = `status-dot status-${{row.status}}`;
            statusTd.appendChild(dot);
            statusTd.appendChild(document.createTextNode(row.status));
            tr.appendChild(statusTd);
            displayCols.forEach(col => {{
                const td = document.createElement('td');
                const cellData = row[col];
                if (cellData.is_diff) td.innerHTML = cellData.val;
                else td.innerText = cellData.val !== null ? cellData.val : '';
                tr.appendChild(td);
            }});
            tbody.appendChild(tr);
        }});
    }}
    function renderTree(node, container, pathName) {{
        const keys = Object.keys(node).filter(k => !k.startsWith('_'));
        keys.forEach(key => {{
            const childNode = node[key];
            const wrapper = document.createElement('div');
            wrapper.className = 'tree-node';
            const label = document.createElement('div');
            label.className = 'tree-label';
            const stats = childNode._stats || {{}};
            const changedCount = stats.changed || 0;
            const badge = changedCount > 0 ? `<span class="badge changed">${{changedCount}}</span>` : '';
            label.innerHTML = `<span>${{key}}</span> ${{badge}}`;
            label.onclick = (e) => {{
                e.stopPropagation();
                const childrenContainer = wrapper.querySelector('.tree-children');
                if (childrenContainer) childrenContainer.classList.toggle('open');
                const ids = collectIds(childNode);
                renderTable(data.filter((r, i) => ids.includes(i)));
                document.getElementById('current-view-title').innerText = `${{pathName}} > ${{key}}`;
                document.querySelectorAll('.tree-label').forEach(l => l.classList.remove('active'));
                label.classList.add('active');
            }};
            wrapper.appendChild(label);
            if (Object.keys(childNode).some(k => !k.startsWith('_'))) {{
                const childrenDiv = document.createElement('div');
                childrenDiv.className = 'tree-children';
                renderTree(childNode, childrenDiv, `${{pathName}} > ${{key}}`);
                wrapper.appendChild(childrenDiv);
            }}
            container.appendChild(wrapper);
        }});
    }}
    function collectIds(node) {{
        let ids = node._ids || [];
        Object.keys(node).forEach(key => {{ if (!key.startsWith('_')) ids = ids.concat(collectIds(node[key])); }});
        return ids;
    }}
    renderTable(data);
    renderTree(tree, document.getElementById('tree-root'), 'Root');
</script>
</body>
</html>
    """
    
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html_content)
    
    log(f"HTML report written to {output_path}", 1)

def main():
    parser = argparse.ArgumentParser(description="CSV/Excel Diff Tool with Excel Output")
    parser.add_argument("--left", required=True, help="Path to left file (CSV or Excel)")
    parser.add_argument("--right", required=True, help="Path to right file (CSV or Excel)")
    parser.add_argument("--mapping", help="Path to mapping Excel file")
    parser.add_argument("--mapping-confirmed", action="store_true", help="Flag to indicate mapping is ready")
    parser.add_argument("--out", default="diff_results.xlsx", help="Output Excel file")
    parser.add_argument("--html", help="Output HTML file (optional)")
    parser.add_argument("--hierarchy", help="Comma-separated list of columns for hierarchy (e.g. 'Channel,Label')")
    parser.add_argument("--max-rows-excel", type=int, default=200000, help="Max rows to export to Excel")
    
    # Header Row Arguments
    parser.add_argument("--header-row", type=int, default=1, help="Header row for both files (default: 1)")
    parser.add_argument("--left-header-row", type=int, help="Header row for left file (overrides --header-row)")
    parser.add_argument("--right-header-row", type=int, help="Header row for right file (overrides --header-row)")
    
    # New Arguments
    parser.add_argument("--sheet", type=str, help="Sheet name (for Excel). Defaults to first sheet.")
    parser.add_argument("--left-sheet", type=str, help="Sheet name for left file.")
    parser.add_argument("--right-sheet", type=str, help="Sheet name for right file.")

    parser.add_argument("--debug", type=int, default=0, help="Debug level (1=Info, 2=Flow, 3=Data)")

    args = parser.parse_args()
    
    global DEBUG_LEVEL
    DEBUG_LEVEL = args.debug

    # Resolve header rows
    left_header = args.left_header_row if args.left_header_row is not None else args.header_row
    right_header = args.right_header_row if args.right_header_row is not None else args.header_row

    # Resolve sheets
    left_sheet = args.left_sheet if args.left_sheet is not None else args.sheet
    right_sheet = args.right_sheet if args.right_sheet is not None else args.sheet

    # 1. Check if mapping is needed
    if not args.mapping:
        if args.mapping_confirmed:
            log("Error: --mapping-confirmed passed but no mapping file specified.", 1)
            sys.exit(1)
        
        mapping_path = "mapping_template.xlsx"
        if os.path.exists(mapping_path):
             log(f"Mapping template {mapping_path} already exists.", 1)
             log("Please edit it and run with --mapping " + mapping_path + " --mapping-confirmed", 1)
             pass
        
        # Always create if not provided?
        create_mapping_template(args.left, args.right, mapping_path, left_header_row=left_header, right_header_row=right_header, left_sheet=left_sheet, right_sheet=right_sheet)
        sys.exit(0)
    else:
        # User provided mapping
        if not args.mapping_confirmed:
            log(f"Mapping file {args.mapping} provided.", 1)
            log("Please add --mapping-confirmed to proceed with diff.", 1)
            
            # Check if template needs creating?
            if not os.path.exists(args.mapping):
                log(f"Mapping file {args.mapping} does not exist. Creating it...", 1)
                create_mapping_template(args.left, args.right, args.mapping, left_header_row=left_header, right_header_row=right_header, left_sheet=left_sheet, right_sheet=right_sheet)
                sys.exit(0)
            else:
                log(f"Using existing default mapping: {args.mapping}", 1)
                args.mapping = args.mapping

    # 3. Read Mapping
    import time
    t0 = time.time()
    mapping_dict, keys, fill_down_cols = read_mapping(args.mapping)
    t1 = time.time()
    log(f"Mapping read time: {t1-t0:.4f}s", 1)
    
    # 4. Compute Diff
    t2 = time.time()
    df_diff = compute_diff(args.left, args.right, mapping_dict, keys, fill_down_cols=fill_down_cols, 
                           left_header_row=left_header, right_header_row=right_header, 
                           left_sheet=left_sheet, right_sheet=right_sheet)
    t3 = time.time()
    log(f"Diff computation time: {t3-t2:.4f}s", 1)
    
    # 5. Write Output
    write_excel_report(df_diff, args.out, mapping_dict, keys, max_rows=args.max_rows_excel)

    # 6. HTML Output
    if args.html:
        hierarchy_cols = [c.strip() for c in args.hierarchy.split(",")] if args.hierarchy else []
        write_html_report(df_diff, args.html, mapping_dict, keys, hierarchy_cols)

if __name__ == "__main__":
    main()

import openpyxl
import sys
import os

def configure_mapping(path):
    print(f"Configuring {path}...")
    if not os.path.exists(path):
        print(f"Error: {path} not found.")
        return

    try:
        # Load workbook (standard load is usually safe for sparse files in openpyxl, 
        # unlike pandas which densifies. If it's still an issue, we'd need streaming read + write, which is complex.
        # But usually openpyxl handles high dimensions with empty cells better than pandas.)
        wb = openpyxl.load_workbook(path)
        if 'Columns' not in wb.sheetnames:
            print("Error: 'Columns' sheet not found.")
            return
        
        ws = wb['Columns']
        
        # Find headers
        header_map = {}
        header_row_idx = None
        
        # Scan first few rows for headers
        for i, row in enumerate(ws.iter_rows(max_row=10, values_only=True)):
            row_lower = [str(c).strip().lower() if c is not None else "" for c in row]
            if 'left_column' in row_lower:
                for idx, val in enumerate(row_lower):
                    if val:
                        header_map[val] = idx
                header_row_idx = i + 1 # 1-based for openpyxl
                break
        
        if header_row_idx is None:
            print("Error: Headers not found.")
            return

        col_left = header_map.get('left_column')
        col_key = header_map.get('is_key')
        col_fill = header_map.get('fill_down')
        
        if col_left is None:
            print("Error: 'left_column' header missing.")
            return

        print("Updating rows...")
        rows_updated = 0
        
        # Iterate rows (skip header)
        # We use iter_rows returning Cell objects to modify them
        for row in ws.iter_rows(min_row=header_row_idx + 1):
             # Check for phantom rows
             if not any(c.value for c in row):
                 continue
                 
             val_left = row[col_left].value
             if val_left is None:
                 continue
                 
             s_left = str(val_left).strip()
             
             # Apply Logic
             if s_left == 'ID' and col_key is not None:
                 row[col_key].value = 'Y'
                 rows_updated += 1
             
             if s_left in ['Category', 'Subcategory'] and col_fill is not None:
                 row[col_fill].value = 'Y'
                 rows_updated += 1
                 
        print(f"Updated {rows_updated} cells.")
        wb.save(path)
        print("Saved.")
        
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    configure_mapping('mapping_template.xlsx')
